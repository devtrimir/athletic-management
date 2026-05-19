"""
Phase 1 extraction: dump every sheet of every workbook in the project root
into inspectable CSV + cell-grid JSON + a structural manifest.

Outputs (under analysis/):
  raw_csv/<workbook>/<sheet>.csv          - plain values, UTF-8, Devanagari preserved
  raw_json/<workbook>/<sheet>.json        - rich cell grid w/ merge/format/formula metadata
  manifest/workbooks.json                 - machine-readable structural manifest
  manifest/workbooks.md                   - human-readable structural manifest
  manifest/headers_inventory.csv          - candidate header cells across all sheets
  manifest/value_samples.json             - per-column samples + cardinality + null%

Usage:
  .venv/bin/python analysis/scripts/extract_workbooks.py
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "analysis"
RAW_CSV = OUT / "raw_csv"
RAW_JSON = OUT / "raw_json"
MANIFEST = OUT / "manifest"

WORKBOOKS = [
    "2019 se 2025 (1).xlsx",
    "2026 (1).xlsx",
    "UP POLICE TEAM PLAYERS DETAILS UPDATED.xlsx",
]

SAFE = re.compile(r"[^A-Za-z0-9._-]+")


def slug(name: str) -> str:
    s = SAFE.sub("_", name.strip()).strip("_")
    return s or "sheet"


def cell_value(v: Any) -> Any:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    return v


def build_merge_index(ws) -> tuple[dict, dict]:
    """Return (anchor_map, member_map). anchor_map[(r,c)] = (rows, cols).
    member_map[(r,c)] = (anchor_r, anchor_c) for non-anchor merged members."""
    anchor: dict[tuple[int, int], tuple[int, int]] = {}
    member: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        ar, ac = rng.min_row, rng.min_col
        rows = rng.max_row - rng.min_row + 1
        cols = rng.max_col - rng.min_col + 1
        anchor[(ar, ac)] = (rows, cols)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (ar, ac):
                    member[(r, c)] = (ar, ac)
    return anchor, member


def real_bounds(ws, declared_max_row: int, declared_max_col: int,
                hard_cap_rows: int = 20000, hard_cap_cols: int = 200) -> tuple[int, int]:
    """Excel often reports max_row/max_col as the sheet's allocated dimension
    (e.g., 1,048,575) even when only a few hundred rows actually contain data.
    Walk from the bottom-right backwards to find the last non-empty cell.
    Hard caps protect against pathological sheets."""
    max_row = min(declared_max_row or 0, hard_cap_rows)
    max_col = min(declared_max_col or 0, hard_cap_cols)
    if max_row == 0 or max_col == 0:
        return 0, 0

    # Find true last non-empty row by scanning from the top using iter_rows
    last_row = 0
    last_col = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col,
                            values_only=True):
        # row index isn't given here; track via counter
        pass
    # Re-iterate with row index
    r_idx = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col,
                            values_only=True):
        r_idx += 1
        # rightmost non-empty in this row
        for ci in range(len(row) - 1, -1, -1):
            v = row[ci]
            if v is not None and (not isinstance(v, str) or v.strip() != ""):
                last_row = r_idx
                if ci + 1 > last_col:
                    last_col = ci + 1
                break
    return last_row, last_col


def extract_sheet(wb_path: Path, sheet_name: str, ws) -> dict:
    anchor, member = build_merge_index(ws)
    hidden_cols = {
        idx for idx, dim in ws.column_dimensions.items() if dim.hidden
    }
    # column_dimensions keys are letters; convert
    hidden_col_idxs = set()
    for letter in hidden_cols:
        try:
            from openpyxl.utils import column_index_from_string

            hidden_col_idxs.add(column_index_from_string(letter))
        except Exception:
            pass
    hidden_rows = {
        idx for idx, dim in ws.row_dimensions.items() if dim.hidden
    }

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    # Trim trailing empties (Excel may report allocated dimension, not used).
    trimmed_row, trimmed_col = real_bounds(ws, max_row, max_col)
    if trimmed_row and trimmed_col:
        max_row, max_col = trimmed_row, trimmed_col

    grid: list[list[dict]] = []
    csv_rows: list[list[Any]] = []
    formula_count = 0

    for r in range(1, max_row + 1):
        json_row = []
        csv_row = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            raw = cell.value
            is_formula = isinstance(raw, str) and raw.startswith("=")
            formula = raw if is_formula else None
            value = None if is_formula else cell_value(raw)
            entry = {
                "v": value,
                "f": formula,
            }
            if (r, c) in anchor:
                rows, cols = anchor[(r, c)]
                entry["merge_anchor"] = True
                entry["merge_rows"] = rows
                entry["merge_cols"] = cols
            elif (r, c) in member:
                ar, ac = member[(r, c)]
                entry["merge_member_of"] = [ar, ac]
            if r in hidden_rows:
                entry["row_hidden"] = True
            if c in hidden_col_idxs:
                entry["col_hidden"] = True
            try:
                font = cell.font
                if font and font.bold:
                    entry["bold"] = True
            except Exception:
                pass
            try:
                fill = cell.fill
                if fill and getattr(fill, "fgColor", None) and fill.fgColor.rgb and fill.fgColor.rgb not in ("00000000",):
                    entry["fill"] = str(fill.fgColor.rgb)
            except Exception:
                pass
            if is_formula:
                formula_count += 1
            json_row.append(entry)

            # For CSV, fill merged members with the anchor's value (common spreadsheet idiom)
            if (r, c) in member:
                ar, ac = member[(r, c)]
                anchor_cell = ws.cell(row=ar, column=ac).value
                csv_row.append(cell_value(anchor_cell))
            else:
                csv_row.append(value if not is_formula else f"={formula}")
        grid.append(json_row)
        csv_rows.append(csv_row)

    wb_slug = slug(wb_path.stem)
    sh_slug = slug(sheet_name)
    csv_dir = RAW_CSV / wb_slug
    json_dir = RAW_JSON / wb_slug
    csv_dir.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)

    csv_path = csv_dir / f"{sh_slug}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerows(csv_rows)

    json_path = json_dir / f"{sh_slug}.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "workbook": wb_path.name,
                "sheet": sheet_name,
                "max_row": max_row,
                "max_col": max_col,
                "grid": grid,
            },
            f,
            ensure_ascii=False,
        )

    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    sample_rows = csv_rows[:20]

    return {
        "sheet": sheet_name,
        "slug": sh_slug,
        "max_row": max_row,
        "max_col": max_col,
        "merged_ranges": merged_ranges,
        "merged_count": len(merged_ranges),
        "hidden_rows": sorted(hidden_rows),
        "hidden_cols": sorted(hidden_col_idxs),
        "formula_count": formula_count,
        "csv_path": str(csv_path.relative_to(ROOT)),
        "json_path": str(json_path.relative_to(ROOT)),
        "sample_rows": sample_rows,
    }


def detect_header_row(sample_rows: list[list[Any]]) -> int | None:
    """Heuristic: first row in the first 10 where >=3 string cells exist and
    >=70% of non-null cells are short strings (<60 chars)."""
    best = None
    best_score = 0
    for i, row in enumerate(sample_rows[:10]):
        non_null = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_null) < 3:
            continue
        strs = [c for c in non_null if isinstance(c, str) and len(c) < 60]
        ratio = len(strs) / len(non_null)
        score = len(strs) * ratio
        if ratio >= 0.7 and score > best_score:
            best_score = score
            best = i
    return best


def harvest_headers(manifest: dict) -> list[dict]:
    """Build headers inventory: workbook, sheet, header_row_index, col, text."""
    rows = []
    for wb in manifest["workbooks"]:
        for sh in wb["sheets"]:
            hdr = sh.get("header_row")
            if hdr is None:
                continue
            header_row = sh["sample_rows"][hdr]
            for ci, val in enumerate(header_row, start=1):
                if val is None:
                    continue
                text = str(val).strip()
                if not text:
                    continue
                rows.append(
                    {
                        "workbook": wb["workbook"],
                        "sheet": sh["sheet"],
                        "header_row": hdr + 1,
                        "col": ci,
                        "col_letter": get_column_letter(ci),
                        "text": text,
                    }
                )
    return rows


def value_samples(manifest: dict, max_samples: int = 10) -> dict:
    """For each (workbook, sheet, column), collect samples + cardinality + null%.
    Re-read CSV for full coverage (sample_rows in manifest is only top 20)."""
    out: dict = {}
    for wb in manifest["workbooks"]:
        out[wb["workbook"]] = {}
        for sh in wb["sheets"]:
            hdr_idx = sh.get("header_row")
            csv_path = ROOT / sh["csv_path"]
            with csv_path.open(encoding="utf-8") as f:
                reader = list(csv.reader(f))
            if not reader:
                out[wb["workbook"]][sh["sheet"]] = {}
                continue
            header = reader[hdr_idx] if hdr_idx is not None and hdr_idx < len(reader) else []
            data = reader[(hdr_idx + 1):] if hdr_idx is not None else reader
            cols: dict[int, dict] = {}
            ncols = max((len(r) for r in reader), default=0)
            for ci in range(ncols):
                values = [
                    (r[ci] if ci < len(r) else "")
                    for r in data
                ]
                non_null = [v for v in values if v is not None and str(v).strip() != ""]
                counter = Counter(non_null)
                cols[ci + 1] = {
                    "header": header[ci] if ci < len(header) else None,
                    "total_rows": len(values),
                    "non_null": len(non_null),
                    "null_pct": round(100.0 * (len(values) - len(non_null)) / len(values), 2) if values else 0.0,
                    "distinct": len(counter),
                    "top_samples": [
                        {"value": v, "count": n}
                        for v, n in counter.most_common(max_samples)
                    ],
                }
            out[wb["workbook"]][sh["sheet"]] = cols
    return out


def render_markdown(manifest: dict) -> str:
    lines = ["# Workbook Structural Manifest", ""]
    for wb in manifest["workbooks"]:
        lines.append(f"## {wb['workbook']}")
        lines.append("")
        lines.append(f"- Sheets: **{len(wb['sheets'])}**")
        lines.append("")
        for sh in wb["sheets"]:
            lines.append(f"### Sheet: `{sh['sheet']}`")
            lines.append("")
            lines.append(f"- Dimensions: {sh['max_row']} rows × {sh['max_col']} cols")
            lines.append(f"- Merged ranges: {sh['merged_count']}")
            lines.append(f"- Hidden rows: {len(sh['hidden_rows'])}  •  Hidden cols: {len(sh['hidden_cols'])}")
            lines.append(f"- Formula cells: {sh['formula_count']}")
            lines.append(f"- Detected header row (0-based): {sh.get('header_row')}")
            lines.append(f"- CSV: `{sh['csv_path']}`")
            lines.append(f"- JSON: `{sh['json_path']}`")
            if sh["merged_ranges"]:
                show = ", ".join(sh["merged_ranges"][:8])
                more = "" if len(sh["merged_ranges"]) <= 8 else f" … (+{len(sh['merged_ranges']) - 8} more)"
                lines.append(f"- Sample merges: {show}{more}")
            lines.append("")
            lines.append("**First 5 rows (post merged-fill):**")
            lines.append("")
            lines.append("```")
            for row in sh["sample_rows"][:5]:
                trimmed = [
                    ("" if v is None else str(v))[:40]
                    for v in row[:12]
                ]
                lines.append(" | ".join(trimmed))
            lines.append("```")
            lines.append("")
    return "\n".join(lines)


def main() -> None:
    OUT.mkdir(exist_ok=True)
    RAW_CSV.mkdir(exist_ok=True)
    RAW_JSON.mkdir(exist_ok=True)
    MANIFEST.mkdir(exist_ok=True)

    manifest = {"workbooks": []}

    for wb_name in WORKBOOKS:
        wb_path = ROOT / wb_name
        if not wb_path.exists():
            print(f"[WARN] missing: {wb_name}")
            continue
        print(f"[INFO] opening {wb_name}")
        wb = load_workbook(wb_path, data_only=True, read_only=False)
        wb_entry = {
            "workbook": wb_name,
            "sheet_names": wb.sheetnames,
            "sheets": [],
        }
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"  - sheet: {sn} ({ws.max_row}x{ws.max_column})")
            sh_entry = extract_sheet(wb_path, sn, ws)
            sh_entry["header_row"] = detect_header_row(sh_entry["sample_rows"])
            wb_entry["sheets"].append(sh_entry)
        manifest["workbooks"].append(wb_entry)

    (MANIFEST / "workbooks.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (MANIFEST / "workbooks.md").write_text(render_markdown(manifest), encoding="utf-8")

    headers = harvest_headers(manifest)
    with (MANIFEST / "headers_inventory.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["workbook", "sheet", "header_row", "col", "col_letter", "text"],
        )
        w.writeheader()
        w.writerows(headers)

    # Distinct header text frequency, sorted
    freq = Counter(h["text"] for h in headers)
    with (MANIFEST / "headers_distinct.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["text", "occurrences"])
        for text, n in freq.most_common():
            w.writerow([text, n])

    samples = value_samples(manifest)
    (MANIFEST / "value_samples.json").write_text(
        json.dumps(samples, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("\n[DONE]")
    print(f"  workbooks analyzed: {len(manifest['workbooks'])}")
    print(f"  total sheets: {sum(len(w['sheets']) for w in manifest['workbooks'])}")
    print(f"  distinct header strings: {len(freq)}")
    print(f"  manifest: {MANIFEST}/workbooks.md")


if __name__ == "__main__":
    main()
